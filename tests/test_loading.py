"""Loading and packaging.

The bugs here are the quiet kind. A multi-sheet workbook that reads its cover
note instead of the responses doesn't raise — it just hands you a two-row
dataset and lets you analyse it. Every test below started as a real failure.
"""

from __future__ import annotations

import ast
import io
import re
import sys
import tomllib
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT))

from hma import loaders  # noqa: E402
from scripts.make_messy_survey import build  # noqa: E402


@pytest.fixture(scope="module")
def messy() -> pd.DataFrame:
    return build()


def _buf(path: Path) -> io.BytesIO:
    return io.BytesIO(path.read_bytes())


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------

def test_plain_xlsx_round_trips(messy, tmp_path):
    path = tmp_path / "plain.xlsx"
    messy.to_excel(path, index=False)
    result = loaders.load(_buf(path), "plain.xlsx")
    assert result.df.shape == messy.shape


def test_multi_sheet_picks_the_data_not_the_cover_note(messy, tmp_path):
    """The bug that started this: pd.read_excel takes sheet 0, silently.

    Survey platforms put a ReadMe or cover sheet first. Reading blind gives
    you a two-row dataset and no indication anything went wrong.
    """
    path = tmp_path / "multi.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame({"note": ["Exported 2026-07-15", "Contact: research@x.org"]}).to_excel(
            writer, sheet_name="ReadMe", index=False
        )
        messy.to_excel(writer, sheet_name="Responses", index=False)
        pd.DataFrame({"code": [1, 2], "label": ["Yes", "No"]}).to_excel(
            writer, sheet_name="Codebook", index=False
        )

    result = loaders.load(_buf(path), "multi.xlsx")
    assert result.sheet == "Responses", f"picked {result.sheet!r} instead of the data"
    assert result.df.shape == messy.shape
    assert result.notes, "picking a sheet silently is the same bug in a new coat"
    assert "Responses" in result.notes[0]


def test_sheets_are_listed_for_the_user_to_override(messy, tmp_path):
    """A heuristic that can't be overridden is just a confident guess."""
    path = tmp_path / "multi.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame({"note": ["cover"]}).to_excel(writer, sheet_name="ReadMe", index=False)
        messy.to_excel(writer, sheet_name="Responses", index=False)

    sheets = loaders.excel_sheets(_buf(path), "multi.xlsx")
    assert {s.name for s in sheets} == {"ReadMe", "Responses"}
    assert sheets[0].name == "Responses", "best candidate should sort first"

    forced = loaders.load(_buf(path), "multi.xlsx", sheet="ReadMe")
    assert forced.sheet == "ReadMe", "explicit sheet choice must win over the heuristic"


def test_title_rows_above_the_header_are_skipped(messy, tmp_path):
    """Excel exports love a title block. Read blind and it becomes the header."""
    path = tmp_path / "titled.xlsx"
    messy.to_excel(path, index=False, startrow=3)
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="MALARIA KAP SURVEY 2026")
    sheet.cell(row=2, column=1, value="Kwara State")
    workbook.save(path)

    result = loaders.load(_buf(path), "titled.xlsx")
    assert list(result.df.columns)[:2] == ["respondent_id", "respondent_name"]
    assert result.df.shape == messy.shape
    assert any("header" in n.lower() for n in result.notes)


def test_header_detection_leaves_a_clean_file_alone(messy, tmp_path):
    """A detector that 'fixes' healthy files is worse than no detector."""
    path = tmp_path / "clean.xlsx"
    messy.to_excel(path, index=False)
    result = loaders.load(_buf(path), "clean.xlsx")
    assert list(result.df.columns) == list(messy.columns)


# --------------------------------------------------------------------------
# CSV / JSON
# --------------------------------------------------------------------------

def test_csv_round_trips(messy, tmp_path):
    path = tmp_path / "d.csv"
    messy.to_csv(path, index=False)
    assert loaders.load(_buf(path), "d.csv").df.shape == messy.shape


def test_json_round_trips(messy, tmp_path):
    path = tmp_path / "d.json"
    messy.to_json(path, orient="records")
    assert loaders.load(_buf(path), "d.json").df.shape[0] == messy.shape[0]


def test_unsupported_extension_says_what_is_supported():
    with pytest.raises(loaders.UnsupportedFile) as err:
        loaders.load(io.BytesIO(b"x"), "notes.docx")
    assert "csv" in str(err.value).lower()


def test_empty_file_fails_clearly_not_cryptically():
    with pytest.raises(loaders.UnsupportedFile):
        loaders.load(io.BytesIO(b""), "empty.csv")


# --------------------------------------------------------------------------
# Packaging — the class of bug that only bites on someone else's machine
# --------------------------------------------------------------------------

def _declared() -> set[str]:
    cfg = tomllib.loads((ROOT / "pyproject.toml").read_text())
    names = cfg["project"]["dependencies"] + cfg["project"]["optional-dependencies"]["dev"]
    return {re.split(r"[><=\[;]", n)[0].strip().lower().replace("-", "_") for n in names}


def test_every_import_is_a_declared_dependency():
    """Undeclared deps work fine here and explode on a fresh `uv sync`.

    matplotlib shipped this way once: charts.py imported it, pyproject never
    mentioned it, and every test passed because the container happened to
    have it installed.
    """
    stdlib = set(sys.stdlib_module_names)
    local = {"hma", "scripts", "app", "tests"}
    aliases = {"pil": "pillow", "markdown_it": "markdown_it_py"}

    files = [
        *(ROOT / "src").rglob("*.py"),
        *(ROOT / "scripts").rglob("*.py"),
        *(ROOT / "tests").rglob("*.py"),
        ROOT / "app.py",
    ]
    imported: set[str] = set()
    for file in files:
        for node in ast.walk(ast.parse(file.read_text())):
            if isinstance(node, ast.Import):
                imported |= {a.name.split(".")[0] for a in node.names}
            elif isinstance(node, ast.ImportFrom) and node.level == 0 and node.module:
                imported.add(node.module.split(".")[0])

    declared = _declared()
    missing = sorted(
        m for m in imported
        if m not in stdlib and m not in local
        and aliases.get(m.lower(), m.lower()) not in declared
    )
    assert not missing, f"imported but not declared in pyproject.toml: {missing}"


@pytest.mark.parametrize("engine, extension", [("openpyxl", ".xlsx"), ("xlrd", ".xls")])
def test_excel_engines_are_installed(engine: str, extension: str):
    """pandas loads these dynamically, so no static check can catch them.

    The uploader offers both extensions. If the engine isn't declared, the
    upload raises a pandas ImportError the user can do nothing about.
    """
    pytest.importorskip(engine, reason=f"{engine} missing — every {extension} upload will fail")
    assert engine.lower().replace("-", "_") in _declared(), (
        f"{engine} is needed for {extension} uploads but isn't in pyproject.toml"
    )


def test_uploader_extensions_match_what_loaders_accepts():
    """The file picker must not advertise formats the loader rejects."""
    app = (ROOT / "app.py").read_text()
    match = re.search(r'type=\[([^\]]+)\]', app)
    assert match, "could not find the uploader's type= list in app.py"
    offered = {e.strip().strip("\"'").lower() for e in match.group(1).split(",")}
    supported = {e.lstrip(".") for e in loaders.SUPPORTED}
    assert offered <= supported, f"uploader offers formats loaders can't read: {offered - supported}"
